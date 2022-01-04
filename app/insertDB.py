from app import db
import requests
from openpyxl import load_workbook  # 엑셀을 불러오기위해 openpyxl 패키지 사용
from kakao_api import generate_location
from models import Vegan, User
from bs4 import BeautifulSoup

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari / 537.36'}


def insert_db():
    wb = load_workbook(filename='VeganMap.xlsx')
    # 엑셀 불러오기
    ws = wb.active
    # 현재 활성화 되어있는 시트 선택! [ 저는 시트가 하나라서 그게 선택됩니다!]

    for row in range(3, ws.max_row):
        try:
            lat_log = generate_location(ws.cell(row, 6).value)
        except Exception as e:
            lat_log = [f"error{e}", f"error{e}"]
        vegan = Vegan(shop=ws.cell(row, 2).value, address=ws.cell(row, 6).value, sector=ws.cell(row, 3).value,
                      menu=ws.cell(row, 7).value, longitude=lat_log[0], latitude=lat_log[1],
                      region=ws.cell(row, 5).value,
                      image=image_scraping(ws.cell(row, 5).value + " " + ws.cell(row, 2).value))

        db.session.add(vegan)
        print(f"add 성공: {ws.cell(row, 2).value}")

    db.session.commit()


def image_scraping(search):
    data = requests.get(
        f'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query={search}',
        headers=headers)
    soup = BeautifulSoup(data.text, 'html.parser')
    if data.status_code == 200:
        try:
            images = soup.select('#place_main_ct')
            if len(images) != 0:
                for image in images:
                    img_url = \
                        image.select_one(
                            'div > section > div > div.top_photo_area.type_v4 > div:nth-child(1) > a > img')[
                            'src']
                    print(f'{search} 성공')
                    vegan = img_url
            else:
                print(f'{search} 실패')
            vegan = f"../static/img/carrot.jpg"

        except Exception as e:
            print(f'{search} 실패')
        vegan = f"../static/img/carrot.jpg"


user = User(user_id="tester",
            password="testpassword",
            email="test@example.com",
            )
db.session.add(user)
db.session.commit()
# db.session.delete(Vegan.query.get(972))

# vegan = Vegan(id=1, shop="502 세컨즈카페", address="서울 노원구 섬밭로 232 우성아파트", sector="양식",
#               menu="502 샐러드파스타(락토,비건가능), 502 양념감자튀김(락토,비건가능)", longitude=37.6366151651829, latitude=127.065668150406,
#               region="노원구",
#               image="https://search.pstatic.net/common/?autoRotate=true&quality=95&size=168x130&src=https%3A%2F%2Fldb-phinf.pstatic.net%2F20180323_17%2F1521777524889tQVNp_JPEG%2F_%25A5%25ED___01.jpg&type=f")
# db.session.add(vegan)

# DB 업데이트 방법 1 = select 후 값을 업데이트 한후 commit
# shop_name = Vegan.query.get(1)
# shop_name.shop = "음식점 이름"

# DB 업데이트 방법 2 = select 없이 업데이트 한 후 Commit
# shop_name = db.session.query(Vegan).filter(Vegan.id == 1).update({'shop': '업데이트 테스트!'})

# # id 값이 1인 가게를 가져오기
# vegan_id = Vegan.query.get(1)
# print(f"vegan_id : {vegan_id.shop}")
#
# # 모든 데이터 가져오기
# vegans = Vegan.query.all()
# vegan_all = " ".join(i.shop for i in vegans)
# print(f"vegan_all : {vegan_all}")
#
# # 가게 이름이 "카페 썬"인 멤버를 검색
# vegans = Vegan.query.filter(Vegan.shop == '카페 썬')
# vegan_equal = " ".join(i.shop for i in vegans)
# print(f"vegan_equal: {vegan_equal}")
#
# # 이름이 "카페 썬"가 아닌 멤버를 검색
#
# vegans = Vegan.query.filter(Vegan.shop != "카페 썬")
# vegan_not_equal = " ".join(i.shop for i in vegans)
# print(f"vegan_not_equal: {vegan_not_equal}")
#
# # 이름이 "카페"와 비슷한 멤버를 검색
# # 검색할 문자열 앞에 % 표시 할 경우 %문자열로 끝나는 단어들
# # 검색할 문자열 뒤에 % 표시 할 경우 문자열%로 끝나는 단어들
# # 검색할 문자열 앞뒤에 % 표시시 문자열 포함하는 단어들 %문자열%
# vegans = Vegan.query.filter(Vegan.shop.like("%카페%"))
# vegan_like = " ".join(i.shop for i in vegans)
# print(f"vegan_like: {vegan_like}")
#
# # 이름이 "카페 썬", "5길반찬"에 포함되는 멤버를 검색
# vegans = db.session.query(Vegan).filter(Vegan.shop.in_(("카페 썬", '5길반찬'))).all()
# vegan_in = " ".join(i.shop for i in vegans)
# print(f"vegan_in: {vegan_in}")
#
# # 이름이 "카페 썬", "5길반찬"에 포함되지 않는 멤버를 검색
#
# vegans = Vegan.query.filter(~Vegan.shop.in_(["카페 썬", '5길반찬']))
# vegan_not_in = " ".join(i.shop for i in vegans)
# print(f"vegan_not_in: {vegan_not_in}")
#
# # 이름이 비어있는 멤버를 검색
#
# vegans = Vegan.query.filter(Vegan.shop == None)
# vegan_is_null = " ".join(i.shop for i in vegans)
# print(f"vegan_is_null: {vegan_is_null}")
#
# # 이름이 비어있지 않은 멤버를 검색
#
# vegans = Vegan.query.filter(Vegan.shop != None)
# vegan_is_not_null = " ".join(i.shop for i in vegans)
# print(f"vegan_is_not_null: {vegan_is_not_null}")
#
# # 이름이 "카페 썬"이며 지역이 마포구인 곳을 검색
#
# vegans = Vegan.query.filter((Vegan.shop == "카페 썬") & (Vegan.region == '마포구'))
# vegan_and = " ".join(i.shop for i in vegans)
# print(f"vegan_and: {vegan_and}")
#
# # 이름이 "카페 썬"이거나 지역이 마포구인 곳을 검색
#
# vegans = Vegan.query.filter((Vegan.shop == "카페 썬") | (Vegan.region == '마포구'))
# vegan_or = " ".join(i.shop for i in vegans)
# print(f"vegan_or: {vegan_or}")
#
# # 지역순으로 정렬하여 검색 (기본 = 오름차순)
#
# vegans = Vegan.query.order_by(Vegan.region)
# vegan_order_by = " ".join(i.region for i in vegans)
# print(f"vegan_order_by: {vegan_order_by}")
#
# # 지역을 내림차순으로 정렬하되, limit_num의 크기만큼 반환 [ 내림차 순을 위해 desc() 추가 ]
# limit_num = 5
# vegans = Vegan.query.order_by(Vegan.region.desc()).limit(limit_num)
# vegan_limit = " ".join(i.region for i in vegans)
# print(f"vegan_limit: {vegan_limit}")
#
# # 나이를 내림차순으로 정렬하되, off_set 크기만큼 앞에서 부터 생략하고 반환
# off_set = 5
# vegans = Vegan.query.order_by(Vegan.region.desc()).offset(off_set)
# vegan_offset = " ".join(i.region for i in vegans)
# print(f"vegan_offset: {vegan_offset}")
#
# # 지역을 내림차순으로 정렬하고 나오는 튜플 수를 반환
#
# vegans = Vegan.query.order_by(Vegan.region.desc()).count()
# vegan_count = str(vegans)
# print(f"vegan_count: {vegan_count}")
